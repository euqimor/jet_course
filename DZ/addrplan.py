import os, glob, sys, time
from re import findall
from threading import Thread
from ipaddress import IPv4Network as Network, IPv4Interface as Interface
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

class AnimationThread(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.daemon = True
        self.isRunning = True
    def run(self):
        while self.isRunning:
            spinAnimation()
    def stop(self):
        self.isRunning = False

def spinAnimation():
    spinner = ['/', '-', '\\', '|']
    for item in spinner:
        sys.stdout.write(item)
        sys.stdout.flush()
        sys.stdout.write('\b')
        time.sleep(0.1)

def getAddr(str):
    expAddr = ('interface +(.+) *\n(.*\n(?!!))* *ip address (([0-9]\.?)*) +(([0-9]\.?)*)')
    result = findall(expAddr,str)
    if result:
        return result
    else:
        return False

def getPath():
    path = os.getcwd()
    print('Default filepath is: '+path)
    tmp = input('Press Enter to use the default filepath or enter the path manually:')
    if tmp != '':
        path = tmp
    return path

def getFileList(path):
    fileList = []
    while len(fileList) == 0:
        fileList = glob.glob(path + '*.txt')
        if len(fileList) == 0:
            print('\nNo txt files found under the specified path.')
            path = getPath()
    return fileList

def getAddressList(fileList):
    print('\nWorking, please wait...\n')
    animation = AnimationThread()
    animation.start()
    addressList = []
    for filePath in fileList:
        with open(filePath) as file:
            data = file.read()
            addresses = getAddr(data)
            if addresses:
                for item in addresses:
                    addressList.append(item)
    animation.stop()
    return addressList

def printAddressList(addressList):
    print('\rThe resulting list is ' + str(len(addressList)) + ' items long.')
    answer = input('Are you sure you want to print it?[Y]') or 'Y'
    while True:
        if answer == 'Y' or answer == 'y':
            i = 0
            print('\n')
            for entry in addressList:
                i+=1
                print('|{: 4}| {: <15} {: <15}|'.format(i,entry[2],entry[4]))
                print('|'+'_'*4+'|'+'_'*32+'|')
            break
        elif answer == 'N' or answer == 'n':
            break
        else:
            answer = input('Please answer Y or N [Y]:') or 'Y'

def createNetworkList(addressList):
    networkList = []
    for address in addressList:
        network = Network(address[2]+'/'+address[4],strict=False)
        if network not in networkList:
            networkList.append(network)
    return networkList

def createInterfaceList(addressList):
    interfaceList = []
    for address in addressList:
        interface = Interface(address[2]+'/'+address[4])
        if interface not in interfaceList:
            interfaceList.append(interface)
    return interfaceList

def sortNetworks(networkList):
    def keyFunc(network):
        return (network.netmask._ip,network.network_address._ip)
    return sorted(networkList,key=keyFunc)

def sortInterfaces(interfaceList):
    def keyFunc(interface):
        return (interface.netmask._ip,interface.ip._ip)
    return sorted(interfaceList,key=keyFunc)

def createAddressPlan(networkList, interfaceList):
    addressPlan = {}
    for network in networkList:
        addressPlan.setdefault(str(network),[])
        for interface in interfaceList:
            if interface in network:
                addressPlan[str(network)].append(str(interface.ip))
    return addressPlan

def printAddressPlan(addressPlan):
    print('\rThe resulting list consists of ' + str(len(addressPlan)) + ' subnets.')
    answer = input('Do you want to print it here?[Y]') or 'Y'
    while True:
        if answer.lower() == 'y':
            i = 0
            print('\nAddress Plan:\n')
            for entry in addressPlan:
                i+=1
                print('|{: 4}| {: <30}|'.format(i,entry))
                for item in addressPlan[entry]:
                    print('|{}| {:>30}|'.format(' '*4,item))
                print('|'+'_'*4+'|'+'_'*31+'|')
            break
        elif answer.lower() == 'n':
            break
        else:
            answer = input('Please answer Y or N [Y]:') or 'Y'

def saveInFile(addressPlan, path=os.getcwd(), filename='AddressPlan'):
    wb = Workbook()
    ws = wb.active
    ws.append(['#','Network','IP Address'])
    i = 2
    for subnet in addressPlan:
        ws.cell(row=i,column=1).value = str(i-1)
        ws.cell(row=i,column=2).value = subnet
        tmpStr = ''
        for address in addressPlan[subnet]:
            tmpStr += address + '\n'
        ws.cell(row=i,column=3).alignment = Alignment(wrap_text=True, horizontal='right')
        ws.cell(row=i,column=3).value = tmpStr[:-1]
        i += 1
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    tab = Table(displayName="Table1", ref='A1:C'+str(i-1))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    os.chdir(path)
    wb.save(filename+'.xlsx')
    wb.close()
    print('Saved in '+str(os.getcwd())+' as '+filename+'.xlsx')

def saveDialogue():
    answer = input('\nWould you like to save the data as an excel spreadsheet?[Y]:') or 'Y'
    while answer.lower() not in ['y','n']:
        answer = input('Please answer Y or N[Y]:') or 'Y'
    if answer.lower() == 'y':
        successful = False
        while not successful:
            print('\nCurrent directory is:'+os.getcwd())
            path = input('Enter path to desired directory or press Enter to save in the current directory:') or os.getcwd()
            try:
                os.chdir(path)
                open('testfile','a').close()
                os.remove('testfile')
                successful = True
            except:
                print('Unable to save in the specified directory')
        successful = False
        while not successful:
            filename = input('\nPlease enter the filename without extension [AddressPlan]:') or 'AddressPlan'
            if '.' not in filename and filename+'.xlsx' not in os.listdir(path):
                try:
                    open(filename+'.xlsx','a').close()
                    os.remove(filename+'.xlsx')
                    successful = True
                except:
                    print('Unable to save under specified filename')
            else:
                print('File with that name already exists or you have entered a name with extension')
        return {'answer':answer.lower(), 'path':path, 'filename':filename}



def runProgram():
    path = getPath()
    fileList = getFileList(path)
    addressList = getAddressList(fileList)
    networkList = createNetworkList(addressList)
    networkList = sortNetworks(networkList)
    interfaceList = createInterfaceList(addressList)
    interfaceList = sortInterfaces(interfaceList)
    addressPlan = createAddressPlan(networkList, interfaceList)
    printAddressPlan(addressPlan)
    savePrompt = saveDialogue()
    if savePrompt['answer'] == 'y':
        saveInFile(addressPlan,savePrompt['path'],savePrompt['filename'])

runProgram()